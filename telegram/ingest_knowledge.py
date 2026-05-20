import os
import uuid
import json
import requests
from pathlib import Path
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

DOCS_DIR = Path("/opt/mollo-knowledge/docs")
PROCESSED_DIR = Path("/opt/mollo-knowledge/processed")
COLLECTION = "mollo_empresa"
QDRANT_URL = "http://127.0.0.1:6333"
EMBED_MODEL = "nomic-embed-text"

client = QdrantClient(url=QDRANT_URL)

def embed(text):
    r = requests.post(
        "http://127.0.0.1:11434/api/embeddings",
        json={"model": EMBED_MODEL, "prompt": text},
        timeout=60
    )
    r.raise_for_status()
    return r.json()["embedding"]

def ensure_collection():
    try:
        client.get_collection(COLLECTION)
    except Exception:
        size = len(embed("test"))
        client.create_collection(
            collection_name=COLLECTION,
            vectors_config=VectorParams(size=size, distance=Distance.COSINE)
        )

def read_txt(path):
    return path.read_text(errors="ignore")

def read_pdf(path):
    reader = PdfReader(str(path))
    text = []
    for i, page in enumerate(reader.pages):
        try:
            text.append(f"\n[Página {i+1}]\n" + (page.extract_text() or ""))
        except Exception:
            pass
    return "\n".join(text)

def read_docx(path):
    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs)

def read_xlsx(path):
    wb = load_workbook(str(path), data_only=True)
    rows = []
    for ws in wb.worksheets:
        rows.append(f"\n[Hoja: {ws.title}]")
        for row in ws.iter_rows(values_only=True):
            values = [str(v) for v in row if v is not None]
            if values:
                rows.append(" | ".join(values))
    return "\n".join(rows)

def read_file(path):
    ext = path.suffix.lower()
    if ext == ".txt":
        return read_txt(path)
    if ext == ".pdf":
        return read_pdf(path)
    if ext == ".docx":
        return read_docx(path)
    if ext == ".xlsx":
        return read_xlsx(path)
    return ""

def chunk_text(text, size=1200, overlap=200):
    chunks = []
    start = 0
    while start < len(text):
        chunk = text[start:start+size].strip()
        if chunk:
            chunks.append(chunk)
        start += size - overlap
    return chunks

def main():
    ensure_collection()
    files = list(DOCS_DIR.glob("*"))

    if not files:
        print("No hay documentos en /opt/mollo-knowledge/docs")
        return

    total = 0

    for path in files:
        if path.is_dir():
            continue

        text = read_file(path)
        if not text.strip():
            print(f"Saltando sin texto: {path.name}")
            continue

        chunks = chunk_text(text)
        points = []

        for idx, chunk in enumerate(chunks):
            vector = embed(chunk)
            points.append(
                PointStruct(
                    id=str(uuid.uuid4()),
                    vector=vector,
                    payload={
                        "source": path.name,
                        "chunk": idx,
                        "text": chunk
                    }
                )
            )

        client.upsert(collection_name=COLLECTION, points=points)
        total += len(points)

        marker = PROCESSED_DIR / f"{path.name}.json"
        marker.write_text(json.dumps({"file": path.name, "chunks": len(points)}))

        print(f"Cargado: {path.name} | chunks: {len(points)}")

    print(f"Memoria actualizada. Total chunks nuevos: {total}")

if __name__ == "__main__":
    main()
